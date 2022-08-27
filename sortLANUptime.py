# This is a module written to sort the weekly LAN uptime data for all locations. The report is exported from the Solarwinds server, and it does the following:
# 1. Get the first and last date of the week.
# 2. Change the locations/sites  name to uppercase to avoid duplication.
# 3. Create a seperate sheet for the week's report.
# 4. Copies the data for each location/site for each day of the week to the new sheet.
# 6. Calculates the average per location/site for the week & overall average for the week for each location/site.
# 7. Use conditional formatting to format each cell according to set service level agreements (SLAs).

#importing the necessary modules
import openpyxl, os
from openpyxl.styles.fills import PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.formatting import Rule
from openpyxl.formatting.rule import ColorScaleRule, CellIsRule, FormulaRule
from openpyxl.styles import Alignment, Font, colors, Border, Side
from datetime import date, timedelta

def sortUptime(exportedReport):
	wb = openpyxl.load_workbook(exportedReport)
	sheet1 = wb['Sheet1']
	#picking the date for the first day of the week
	day_one = sheet1.cell(row = 2, column=1).value
	locations = []
	#create new sheet for the week's report
	wb.create_sheet(index=2,title='Uptime for the Week')
	report_sheet = wb['Uptime for the Week']
	#heading for each column
	report_sheet.cell(row=1,column=1,value='S/N')
	report_sheet.cell(row=1,column=2,value='Location/Offsite')
	report_sheet.cell(row=1, column=3).value = str(day_one.strftime('%B %d')) #day one
	report_sheet.cell(row=1, column=4).value = str((day_one + timedelta(days=1)).strftime('%B %d')) #day two
	report_sheet.cell(row=1, column=5).value = str((day_one + timedelta(days=2)).strftime('%B %d')) #day three
	report_sheet.cell(row=1, column=6).value = str((day_one + timedelta(days=3)).strftime('%B %d')) #day four
	report_sheet.cell(row=1, column=7).value = str((day_one + timedelta(days=4)).strftime('%B %d')) #day five
	report_sheet.cell(row=1, column=8, value='Wk/Avg.')	
	for row in range(2, sheet1.max_row + 1):
		#change the location name to uppercase
		sheet1.cell(row=row, column=2).value = sheet1.cell(row=row, column=2).value.upper()
		if sheet1.cell(row=row,column=2).value not in locations:
			locations.append(sheet1.cell(row=row, column=2).value)
		#get the site for each day of the week
		for idx, location in enumerate(locations):
			#serial number for the report sheet
			report_sheet.cell(row=idx + 2, column=1, value=idx + 1)
			#copy the name of each location/offsite to the report sheet
			report_sheet.cell(row=idx + 2, column=2, value=location)
			#copy value for each days of the week for location/offsite
			if sheet1.cell(row=row, column=1).value == day_one and sheet1.cell(row=row, column=2).value == location:
				report_sheet.cell(row=idx + 2, column=3).value = sheet1.cell(row=row, column=4).value
			#second day of the week
			if sheet1.cell(row=row, column=1).value == day_one + timedelta(days=1) and sheet1.cell(row=row, column=2).value == location:
				report_sheet.cell(row=idx + 2, column=4).value = sheet1.cell(row=row, column=4).value
			#third day of the week
			if sheet1.cell(row=row, column=1).value == day_one + timedelta(days=2) and sheet1.cell(row=row, column=2).value == location:
				report_sheet.cell(row=idx + 2, column=5).value = sheet1.cell(row=row, column=4).value
			#fourth day of the week
			if sheet1.cell(row=row, column=1).value == day_one + timedelta(days=3) and sheet1.cell(row=row, column=2).value == location:
				report_sheet.cell(row=idx + 2, column=6).value = sheet1.cell(row=row, column=4).value
			#fifth day of the week
			if sheet1.cell(row=row, column=1).value == day_one + timedelta(days=4) and sheet1.cell(row=row, column=2).value == location:
				report_sheet.cell(row=idx + 2, column=7).value = sheet1.cell(row=row, column=4).value
	#calculate the average for each location/offsite for the week
	for row in range(2, report_sheet.max_row + 1):
		location_average = '=AVERAGE('+get_column_letter(report_sheet.min_column + 2)+str(row)+':'\
							+get_column_letter(report_sheet.max_column - 1)+str(row)+')'
		report_sheet.cell(row=row, column=8, value=location_average).number_format = '00.00'
		#calculate the overall average across the locations/offsites for the week
	overall_average = '=AVERAGE('+get_column_letter(report_sheet.max_column)+str(report_sheet.min_row + 1)+':'\
						+get_column_letter(report_sheet.max_column)+str(report_sheet.max_row)+')'
	report_sheet.cell(row=report_sheet.max_row, column=2, value='AVERAGE')
	report_sheet.cell(row=report_sheet.max_row + 1, column=8, value=overall_average).number_format = '00.00'
	#style for the sheets
    #colors
	grey_fill = PatternFill(start_color='808080', end_color='808080', fill_type='solid')
	purple_fill = PatternFill(start_color='B1A0C7', end_color='B1A0C7', fill_type='solid')
	green_fill = PatternFill(start_color='00B050', end_color='00B050', fill_type='solid')
	lightGreen_fill = PatternFill(start_color='92D050', end_color='92D050', fill_type='solid')
	red_fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
	orange_fill = PatternFill(start_color='FFC000', end_color='FFC000', fill_type='solid')
	cell_border = Border(left=Side(style='medium'), right=Side(style='medium'), top=Side(style='medium'), bottom=Side(style='medium'))
	for row in range(report_sheet.min_row,report_sheet.max_row + 1):
		for col in range (report_sheet.min_column,report_sheet.max_column + 1):
			report_sheet.cell(row=row,column=col).border = cell_border
			report_sheet.cell(row=row,column=col).font = Font(name='Calibri', size = 10, bold=True)
			report_sheet.cell(row=row,column=col).alignment = Alignment(horizontal='center', vertical='center')
			report_sheet.cell(row=row, column=1).fill = grey_fill #color fill for column one
			report_sheet.cell(row=row, column=2).fill = grey_fill #color fill for column two
			report_sheet.cell(row=1, column=col).fill = purple_fill #color fill for row one
			cells = get_column_letter(report_sheet.min_column + 2)+str(row)+':'+get_column_letter(report_sheet.max_column)+str(row)
            #conditional formatting
			report_sheet.conditional_formatting.add(cells, CellIsRule(operator='equal', formula=['100'], stopIfTrue=False, fill=green_fill)) #Exceed Expectation (EE)
			report_sheet.conditional_formatting.add(cells, CellIsRule(operator='equal', formula=['EE'], stopIfTrue=False, fill=green_fill)) #Exceed Expectation (EE)
			report_sheet.conditional_formatting.add(cells, CellIsRule(operator='between', formula=['99.999999', '98.5'], stopIfTrue=False, fill=lightGreen_fill)) #Met Expectation (ME)
			report_sheet.conditional_formatting.add(cells, CellIsRule(operator='between', formula=['98.4999999', '95'], stopIfTrue=False, fill=orange_fill)) #Near Expectation (NE)
			report_sheet.conditional_formatting.add(cells, CellIsRule(operator='lessThan', formula=['95'], stopIfTrue=False, fill=red_fill)) #Below Expectation (BE)\
		report_sheet.cell(row=row,column=1).alignment = Alignment(horizontal='left', vertical='center')
	#delete sheet1 before saving.
	wb.remove(sheet1)
    #storing the date the report was sorted and first day of the week in a variable
	date_sorted = date.today().strftime('%B %d %Y')
	start_date = (date.today() - timedelta(days=4)).strftime('%B %d %Y')
    #saving to the new excel sheet
	wb.save('Location Availability Report - ' + str(start_date).upper() + '- ' + str(date_sorted).upper() + '.xlsx')
	wb.close()
    #return saved file
	return('Location Availability Report - ' + str(start_date).upper() + '- ' + str(date_sorted).upper() + '.xlsx')
