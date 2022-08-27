# This is a module written to sort the weekly uptime data for service providers. The report is exported from the Solarwinds server, and it does the following:
# 1. Get the first and last date of the week.
# 2. Change the providers name to uppercase to avoid duplication.
# 3. Creates a seperate sheet for each provider, with the title as name of provider.
# 4. Store the location/site for each provider in a nested dictionary.
# 5. Copies the data for each provider for each day of the week on their respective sheets.
# 6. Calculates the average per site for the week & overall average for the week for each provider.
# 7. Use conditional formatting to format each cell according to set service level agreements (SLAs).

#importing the necessary modules
import openpyxl, os
from openpyxl.formula.translate import Translator
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.styles.fills import PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.formatting import Rule
from openpyxl.formatting.rule import ColorScaleRule, CellIsRule, FormulaRule
from openpyxl.styles import Alignment, Font, colors, Border, Side
from datetime import date, timedelta
#style for the sheets
#colors
grey_fill = PatternFill(start_color='808080', end_color='808080', fill_type='solid')
purple_fill = PatternFill(start_color='B1A0C7', end_color='B1A0C7', fill_type='solid')
green_fill = PatternFill(start_color='00B050', end_color='00B050', fill_type='solid')
lightGreen_fill = PatternFill(start_color='92D050', end_color='92D050', fill_type='solid')
red_fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
orange_fill = PatternFill(start_color='FFC000', end_color='FFC000', fill_type='solid')
#yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
#white_fill = PatternFill(start_color='FFFFFF', end_color='FFFF00', fill_type='solid')
cell_border = Border(left=Side(style='medium'), right=Side(style='medium'), top=Side(style='medium'), bottom=Side(style='medium'))
def sortUptime(exportedReport):
    wb = openpyxl.load_workbook(exportedReport)
    sheet1 = wb['Sheet1']
    #picking the date for the first day of the week
    day_one = sheet1.cell(row = 2, column=1).value
    #getting the list of providers.
    providers = []
    for row in range(2, sheet1.max_row + 1):
        #change provider names to uppercase
        sheet1.cell(row = row, column = 3).value = sheet1.cell(row = row, column = 3).value.upper()
        if sheet1.cell(row = row, column = 3).value not in providers:
            providers.append(sheet1.cell(row = row, column = 3).value)
    for idx, provider in enumerate(providers):
        # create a sheet for each provider
        wb.create_sheet(index=idx + 2, title=provider)
        #get the site for each provider for the week.
        provider_sheet = wb[provider]
        #heading for each column
        provider_sheet.cell(row=1, column=1, value='Site')
        provider_sheet.cell(row=1, column=2, value='Provider')
        provider_sheet.cell(row=1, column=3).value = str(day_one.strftime('%B %d')) #day one 
        provider_sheet.cell(row=1, column=4).value = str((day_one + timedelta(days=1)).strftime('%B %d')) #day two
        provider_sheet.cell(row=1, column=5).value = str((day_one + timedelta(days=2)).strftime('%B %d')) #day three
        provider_sheet.cell(row=1, column=6).value = str((day_one + timedelta(days=3)).strftime('%B %d')) #day four
        provider_sheet.cell(row=1, column=7).value = str((day_one + timedelta(days=4)).strftime('%B %d')) #day five
        provider_sheet.cell(row=1, column=8, value='Wk/Avg.')
        #declaring nested dictionary to store the sites for each provider e.g {Provider1:{'Site A': '0'}}
        provider_sites = {provider: {}}
        for provider, value in provider_sites.items():
            for row in range(2, sheet1.max_row + 1):
                if sheet1.cell(row = row, column=3).value == provider and sheet1.cell(row = row, column=2).value not in value:
                    value.setdefault(sheet1.cell(row = row, column=2).value, 0)
            #copying the data for each provider
            sites = list(value.keys())
            #copying the value for each site or location or branch
            for row in range(2, sheet1.max_row + 1):
                for rowNum, site in enumerate(sites):
                    #copying the site for the week to each provider sheet
                    provider_sheet.cell(row=rowNum + 2, column=1).value = site
                    provider_sheet.cell(row=rowNum + 2, column=2).value = provider
                    #copying value for the first day of the week.
                    if sheet1.cell(row = row, column = 1).value == day_one and sheet1.cell(row = row, column = 2).value == site:
                        provider_sheet.cell(row=rowNum + 2, column=3).value = sheet1.cell(row = row, column = 4).value   
                    #copying value for the second day of the week.
                    if sheet1.cell(row = row, column = 1).value == day_one + timedelta(days=1) and sheet1.cell(row = row, column = 2).value == site:
                        provider_sheet.cell(row=rowNum + 2, column=4).value = sheet1.cell(row = row, column = 4).value      
                    #copying value for the third day of the week.
                    if sheet1.cell(row = row, column = 1).value == day_one + timedelta(days=2) and sheet1.cell(row = row, column = 2).value == site:
                        provider_sheet.cell(row=rowNum + 2, column=5).value = sheet1.cell(row = row, column = 4).value  
                    #copying value for the fourth day of the week.
                    if sheet1.cell(row = row, column = 1).value == day_one + timedelta(days=3) and sheet1.cell(row = row, column = 2).value == site:
                        provider_sheet.cell(row=rowNum + 2, column=6).value = sheet1.cell(row = row, column = 4).value  
                    #copying value for the last day of the week.
                    if sheet1.cell(row = row, column = 1).value == day_one + timedelta(days=4) and sheet1.cell(row = row, column = 2).value == site:
                        provider_sheet.cell(row=rowNum + 2, column=7).value = sheet1.cell(row = row, column = 4).value   
        #calculate the average per site for the week
        for row in range(2, provider_sheet.max_row + 1):
            site_average = '=AVERAGE('+get_column_letter(provider_sheet.min_column + 2)+str(row)+':'\
                            +get_column_letter(provider_sheet.max_column - 1)+str(row)+')'
            provider_sheet.cell(row=row, column=8, value=site_average).number_format = '00.00'
        #calculate final average for each provider for the week
        week_average = '=AVERAGE('+get_column_letter(provider_sheet.max_column)+str(provider_sheet.min_row + 1)+':'\
                        +get_column_letter(provider_sheet.max_column)+str(provider_sheet.max_row)+')'
        provider_sheet.cell(row=provider_sheet.max_row + 1, column=8, value=week_average).number_format = '00.00'
        provider_sheet.cell(row=provider_sheet.max_row, column=1, value='AVERAGE')
        for row in range(provider_sheet.min_row, provider_sheet.max_row + 1):
            for col in range (provider_sheet.min_column, provider_sheet.max_column + 1):
                provider_sheet.cell(row=row,column=col).border = cell_border
                provider_sheet.cell(row=row,column=col).font = Font(name='Calibri', size = 10, bold=True)
                provider_sheet.cell(row=row,column=col).alignment = Alignment(horizontal='center', vertical='center')
                provider_sheet.cell(row=row, column=1).fill = grey_fill #color fill for column one
                provider_sheet.cell(row=1, column=col).fill = purple_fill #color fill for row one
                cells = get_column_letter(provider_sheet.min_column + 2)+str(row)+':'+get_column_letter(provider_sheet.max_column)+str(row)
                #conditional formatting
                provider_sheet.conditional_formatting.add(cells, CellIsRule(operator='equal', formula=['100'], stopIfTrue=False, fill=green_fill)) #Exceed Expectation (EE)
                provider_sheet.conditional_formatting.add(cells, CellIsRule(operator='equal', formula=['EE'], stopIfTrue=False, fill=green_fill)) #Exceed Expectation (EE)
                provider_sheet.conditional_formatting.add(cells, CellIsRule(operator='between', formula=['99.999999', '98.5'], stopIfTrue=False, fill=lightGreen_fill)) #Met Expectation (ME)
                provider_sheet.conditional_formatting.add(cells, CellIsRule(operator='between', formula=['98.4999999', '95'], stopIfTrue=False, fill=orange_fill)) #Near Expectation (NE)
                provider_sheet.conditional_formatting.add(cells, CellIsRule(operator='lessThan', formula=['95'], stopIfTrue=False, fill=red_fill)) #Below Expectation (BE)
            provider_sheet.cell(row=row,column=1).alignment = Alignment(horizontal='left', vertical='center')      
    #delete sheet1 before saving.
    wb.remove(sheet1)
    #storing the date the report was sorted and first day of the week in a variable
    date_sorted = date.today().strftime('%B %d %Y')
    start_date = (date.today() - timedelta(days=4)).strftime('%B %d %Y')
    #saving to the new excel sheet
    #wb.save(os.environ['userprofile'] + '\\Documents\\WeeklyReport\\' + 'Providers Uptime Report - ' + str(start_date).upper() + '- ' + str(date_sorted).upper() + '.xlsx')
    wb.save('Providers Uptime Report - ' + str(start_date).upper() + '- ' + str(date_sorted).upper() + '.xlsx')
    wb.close()
    #return saved file
    return('Providers Uptime Report - ' + str(start_date).upper() + '- ' + str(date_sorted).upper() + '.xlsx')


