#importing the necessary modules
import openpyxl, os
from openpyxl.formula.translate import Translator
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.styles.fills import PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.formatting import Rule
from openpyxl.formatting.rule import ColorScaleRule, CellIsRule, FormulaRule
from openpyxl.styles import Alignment, Font, colors, Border, Side
from datetime import timedelta
def sort_report(generated_report):
    wb = openpyxl.load_workbook(generated_report)
    sheet1 = wb['Sheet1']
    #picking the date for the first day of the week
    day_one = sheet1.cell(row = 2, column=1).value
    #getting the list of vendors.
    vendors = []
    for row in range(2, sheet1.max_row + 1):
        if sheet1.cell(row = row, column = 3).value not in vendors:
            vendors.append(sheet1.cell(row = row, column = 3).value)
    #create summary sheet for vendors.
    wb.create_sheet(index=2, title='vendors summary')
    summary_sheet = wb['vendors summary']
    #create new sheets for each vendor
    for idx, vendor in enumerate(vendors):
        wb.create_sheet(index=idx + 3, title=vendor)
    #vendor summary sheet data
        summary_sheet.cell(row=1, column=1, value='S/N')
        summary_sheet.cell(row=1, column=2, value='vendor')
        summary_sheet.cell(row=1, column=3, value='% Avg for the week')
        summary_sheet.cell(row=1, column=4, value='Grade')
        summary_sheet.cell(row=idx+2, column=1, value=idx + 1)
        summary_sheet.cell(row=idx+2, column=2, value=vendor)
    #get the site for each vendor for the week.
    for idx, vendor in enumerate(vendors):
        vendor_sheet = wb[vendor]
        #heading for each column
        vendor_sheet.cell(row=1, column=1, value='Site')
        vendor_sheet.cell(row=1, column=2, value='vendor')
        vendor_sheet.cell(row=1, column=3).value = str(day_one.strftime('%B %d')) #day one 
        vendor_sheet.cell(row=1, column=4).value = str((day_one + timedelta(days=1)).strftime('%B %d')) #day two
        vendor_sheet.cell(row=1, column=5).value = str((day_one + timedelta(days=2)).strftime('%B %d')) #day three
        vendor_sheet.cell(row=1, column=6).value = str((day_one + timedelta(days=3)).strftime('%B %d')) #day four
        vendor_sheet.cell(row=1, column=7).value = str((day_one + timedelta(days=4)).strftime('%B %d')) #day five
        vendor_sheet.cell(row=1, column=8, value='Wk/Avg.')
        #declaring nested dictionary to store the sites for each vendor e.g {vendor1:{'Site A': '0'}}
        vendor_sites = {vendor: {}}
        for vendor, value in vendor_sites.items():
            for row in range(2, sheet1.max_row + 1):
                if sheet1.cell(row = row, column=3).value == vendor and sheet1.cell(row = row, column=2).value not in value:
                    value.setdefault(sheet1.cell(row = row, column=2).value, 0)
        #copying data
        for vendor, value in vendor_sites.items():
            #copying the data for each vendor
            sites = list(value.keys())
            for rowNum, site in enumerate(sites):
                #copying the site for the week to each vendor sheet
                vendor_sheet.cell(row=rowNum + 2, column=1).value = site
                vendor_sheet.cell(row=rowNum + 2, column=2).value = vendor
            #copying the value for each sites 
            for row in range(2, sheet1.max_row + 1):
                for rowNum, site in enumerate(sites):
                    #copying value for the first day of the week.
                    if sheet1.cell(row = row, column = 1).value == day_one and sheet1.cell(row = row, column = 2).value == site:
                        vendor_sheet.cell(row=rowNum + 2, column=3).value = sheet1.cell(row = row, column = 4).value   
                    #copying value for the second day of the week.
                    if sheet1.cell(row = row, column = 1).value == day_one + timedelta(days=1) and sheet1.cell(row = row, column = 2).value == site:
                        vendor_sheet.cell(row=rowNum + 2, column=4).value = sheet1.cell(row = row, column = 4).value      
                    #copying value for the third day of the week.
                    if sheet1.cell(row = row, column = 1).value == day_one + timedelta(days=2) and sheet1.cell(row = row, column = 2).value == site:
                        vendor_sheet.cell(row=rowNum + 2, column=5).value = sheet1.cell(row = row, column = 4).value  
                    #copying value for the fourth day of the week.
                    if sheet1.cell(row = row, column = 1).value == day_one + timedelta(days=3) and sheet1.cell(row = row, column = 2).value == site:
                        vendor_sheet.cell(row=rowNum + 2, column=6).value = sheet1.cell(row = row, column = 4).value  
                    #copying value for the last day of the week.
                    if sheet1.cell(row = row, column = 1).value == day_one + timedelta(days=4) and sheet1.cell(row = row, column = 2).value == site:
                        vendor_sheet.cell(row=rowNum + 2, column=7).value = sheet1.cell(row = row, column = 4).value   
        #calculate the average per site for the week
        for row in range(2, vendor_sheet.max_row + 1):
            site_average = '=AVERAGE('+get_column_letter(vendor_sheet.min_column + 2)+str(row)+':'\
                            +get_column_letter(vendor_sheet.max_column - 1)+str(row)+')'
            vendor_sheet.cell(row=row, column=8, value=site_average).number_format = '00.00'
        #calculate final average for each vendor for the week
        week_average = '=AVERAGE('+get_column_letter(vendor_sheet.max_column)+str(vendor_sheet.min_row + 1)+':'\
                        +get_column_letter(vendor_sheet.max_column)+str(vendor_sheet.max_row)+')'
        vendor_sheet.cell(row=vendor_sheet.max_row + 1, column=8, value=week_average).number_format = '00.00'
        vendor_sheet.cell(row=vendor_sheet.max_row, column=1, value='AVERAGE')
        #style for the sheets
        #colors
        grey_fill = PatternFill(start_color='808080', end_color='808080', fill_type='solid')
        purple_fill = PatternFill(start_color='B1A0C7', end_color='B1A0C7', fill_type='solid')
        green_fill = PatternFill(start_color='00B050', end_color='00B050', fill_type='solid')
        lightGreen_fill = PatternFill(start_color='92D050', end_color='92D050', fill_type='solid')
        red_fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
        orange_fill = PatternFill(start_color='FFC000', end_color='FFC000', fill_type='solid')
        #yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
        #white_fill = PatternFill(start_color='FFFFFF', end_color='FFFF00', fill_type='solid')
        cell_border = Border(left=Side(style='medium'), right=Side(style='medium'), top=Side(style='medium'), bottom=Side(style='medium'))
        for row in range(vendor_sheet.min_row, vendor_sheet.max_row + 1):
            for col in range (vendor_sheet.min_column, vendor_sheet.max_column + 1):
                vendor_sheet.cell(row=row,column=col).border = cell_border
                vendor_sheet.cell(row=row,column=col).font = Font(name='Calibri', size = 10, bold=True)
                vendor_sheet.cell(row=row,column=col).alignment = Alignment(horizontal='center', vertical='center')
                vendor_sheet.cell(row=row, column=1).fill = grey_fill #color fill for column one
                vendor_sheet.cell(row=1, column=col).fill = purple_fill #color fill for row one
                cells = get_column_letter(vendor_sheet.min_column + 2)+str(row)+':'+get_column_letter(vendor_sheet.max_column)+str(row)
                #conditional formatting
                vendor_sheet.conditional_formatting.add(cells, CellIsRule(operator='equal', formula=['100'], stopIfTrue=False, fill=green_fill)) 
                vendor_sheet.conditional_formatting.add(cells, CellIsRule(operator='between', formula=['99.999999', '98.5'], stopIfTrue=False, fill=lightGreen_fill)) 
                vendor_sheet.conditional_formatting.add(cells, CellIsRule(operator='between', formula=['98.4999999', '95'], stopIfTrue=False, fill=orange_fill)) 
                vendor_sheet.conditional_formatting.add(cells, CellIsRule(operator='lessThan', formula=['95'], stopIfTrue=False, fill=red_fill)) 
            vendor_sheet.cell(row=row,column=1).alignment = Alignment(horizontal='left', vertical='center')      
    #storing the time the report was sorted in a variable
    sorted_date = datetime.now().strftime('%Y_%m_%d_%I_%M_%S_%p')
    #saving to the new excel sheet
    wb.save(os.environ['userprofile'] + '\\Documents\\Weekly Report\\' + 'Weekly_report_sorted_on_' + sorted_date + '.xlsx')
    wb.close()
    #return saved file
    return(os.environ['userprofile'] + '\\Documents\\Weekly Report\\' + 'Weekly_report_sorted_on_' + sorted_date + '.xlsx')
	
	
	